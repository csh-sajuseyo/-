import sys
import os
import shutil
import atexit
import subprocess
from openpyxl import load_workbook
from PyQt5.QtWidgets import (
    QTableWidget, QTableWidgetItem, QDialog, QVBoxLayout as QVBoxLayout2,
    QApplication, QWidget, QLabel, QPushButton, QVBoxLayout,
    QHBoxLayout, QMessageBox, QFileDialog, QFrame
)
from PyQt5.QtGui import QFont, QCursor
from PyQt5.QtCore import Qt, QTimer
from sheet_api import get_today_pending_requests
from edu_ocr_reader import capture_and_extract_text
from ocr_text_parser import extract_items_with_amount

class SajulgeyoApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("사줄게요 v1.0")
        self.setFixedSize(440, 570)
        self.setStyleSheet("background-color: #f7f9fb;")

        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignTop)

        self.title = QLabel("🛒 사줄게요 v1.0 - 자동구매 도우미")
        self.title.setFont(QFont("맑은 고딕", 14, QFont.Bold))
        self.title.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.title)

        self.request_count_label = QLabel("오늘 요청 건수: ?")
        self.request_count_label.setFont(QFont("맑은 고딕", 12))
        self.request_count_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.request_count_label)

        self.contact_status = self.make_status_label()
        self.mall_status = self.make_status_label()
        self.sheet_status = self.make_status_label()
        self.edu_status = self.make_status_label()
        self.shot_status = self.make_status_label()

        layout.addLayout(self.create_status_row("1.   📇 연락처 불러오기", self.load_contact, self.contact_status, "#e0f3ff", "#d0eaff"))
        layout.addLayout(self.create_status_row("2.   🔗 쇼핑몰 연결 시작하기", self.connect_shopping_mall, self.mall_status, "#f0e7ff", "#e2d6ff"))
        layout.addLayout(self.create_status_row("3.   📂 구매요청서 불러오기", self.load_sheet, self.sheet_status, "#fff9e0", "#fff1c2"))
        layout.addLayout(self.create_status_row("4.   📑 에듀파인 결재 확인 열기", self.run_edu_ocr, self.edu_status, "#fff2e0", "#ffe5c2"))
        layout.addLayout(self.create_status_row("5.   🖼️ 구매 품목 스크린샷 확인", self.dummy_action, self.shot_status, "#e0fff4", "#c2ffe8"))

        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        line.setStyleSheet("color: #ccc; margin: 15px 0;")
        layout.addWidget(line)

        self.run_btn = QPushButton("🚀 물품 자동구매 시작하기")
        self.run_btn.setFont(QFont("맑은 고딕", 11, QFont.Bold))
        self.run_btn.setCursor(QCursor(Qt.PointingHandCursor))
        self.run_btn.setStyleSheet(
            "QPushButton {"
            " background-color: #28a745;"
            " color: white;"
            " padding: 14px;"
            " border-radius: 6px;"
            "}"
            "QPushButton:hover {"
            " background-color: #218838;"
            "}"
            "QPushButton:pressed {"
            " background-color: #1e7e34;"
            "}"
        )
        layout.addWidget(self.run_btn)

        self.setLayout(layout)
        QTimer.singleShot(0, self.center_on_screen)
        self.check_contact_file()

    def make_status_label(self):
        label = QLabel("❔")
        label.setFont(QFont("맑은 고딕", 14))
        label.setFixedWidth(30)
        label.setAlignment(Qt.AlignCenter)
        return label

    def create_status_row(self, text, func, status_label, base_color, hover_color):
        row = QHBoxLayout()
        btn = QPushButton(text)
        btn.setCursor(QCursor(Qt.PointingHandCursor))
        btn.setFont(QFont("맑은 고딕", 10))
        btn.setStyleSheet(
            f"QPushButton {{"
            f" padding: 8px;"
            f" background-color: {base_color};"
            f" border: 1px solid #aaa;"
            f" border-radius: 4px;"
            f" text-align: left;"
            f"}}"
            f"QPushButton:hover {{"
            f" background-color: {hover_color};"
            f"}}"
        )
        btn.clicked.connect(func)
        row.addWidget(btn)
        row.addWidget(status_label)
        return row

    def center_on_screen(self):
        screen = QApplication.primaryScreen().availableGeometry()
        self.move(
            screen.center().x() - self.width() // 2,
            screen.center().y() - self.height() // 2
        )
    def run_edu_ocr(self):
        success, ocr_text = capture_and_extract_text()
        if success:
            try:
                items = extract_items_with_amount(ocr_text)
                if items:
                    dialog = QDialog(self)
                    dialog.setWindowTitle("OCR 품목 인식 결과")
                    dialog.resize(420, 400)
                    table = QTableWidget()
                    table.setColumnCount(2)
                    table.setHorizontalHeaderLabels(["품목명", "금액"])
                    table.setRowCount(len(items))
                    for row, (name, price) in enumerate(items):
                        table.setItem(row, 0, QTableWidgetItem(name))
                        table.setItem(row, 1, QTableWidgetItem(f"{price} 원"))
                    layout = QVBoxLayout2()
                    layout.addWidget(table)
                    dialog.setLayout(layout)
                    dialog.exec_()
                    self.edu_status.setText("✅")
                else:
                    QMessageBox.information(self, "OCR 결과", "📭 인식된 품목이 없습니다.")
                    self.edu_status.setText("❌")
            except Exception as e:
                QMessageBox.critical(self, "후처리 오류", str(e))
                self.edu_status.setText("❌")
        else:
            QMessageBox.critical(self, "OCR 실패", ocr_text)
            self.edu_status.setText("❌")

    def dummy_action(self):
        QMessageBox.information(self, "준비 중", "해당 기능은 곧 지원됩니다!")

    def connect_shopping_mall(self):
        chrome_path = "C:/Program Files/Google/Chrome/Application/chrome.exe"
        user_data = os.path.join(os.getcwd(), "UserData")
        url = "https://login.coupang.com"
        command = [
            chrome_path,
            "--remote-debugging-port=9222",
            f"--user-data-dir={user_data}",
            url
        ]
        try:
            subprocess.Popen(command)
            self.mall_status.setText("✅")
        except Exception as e:
            QMessageBox.critical(self, "실행 오류", f"크롬 실행 실패:\n{e}")
            self.mall_status.setText("❌")

    def load_contact(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "연락처 엑셀파일 선택", "", "Excel Files (*.xlsx)")
        if file_path:
            try:
                dest_path = os.path.join("temp", "contacts.xlsx")
                shutil.copy(file_path, dest_path)
                self.check_contact_file()
            except Exception as e:
                QMessageBox.critical(self, "오류", f"연락처 저장 실패:\n{str(e)}")

    def check_contact_file(self):
        try:
            path = os.path.join("temp", "contacts.xlsx")
            if os.path.exists(path):
                wb = load_workbook(path)
                sheet = wb.active
                rows = list(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True))
                if len(rows) > 0:
                    self.contact_status.setText("✅")
                    return
            self.contact_status.setText("❌")
        except:
            self.contact_status.setText("❌")

    def load_sheet(self):
        try:
            rows = get_today_pending_requests()
            count = len(rows)
            self.request_count_label.setText(f"오늘 요청 건수: {count}")
            self.sheet_status.setText("✅" if count > 0 else "❌")
            print("🔍 오늘 요청 목록:")
            for row in rows:
                print(row)
        except Exception as e:
            QMessageBox.critical(self, "오류", f"구매요청서 불러오기 실패:\n{str(e)}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    CUSTOM_TEMP_DIR = os.path.join(BASE_DIR, "temp")
    os.makedirs(CUSTOM_TEMP_DIR, exist_ok=True)
    LOCK_FILE_PATH = os.path.join(CUSTOM_TEMP_DIR, "sajulgeyo.lock")
    if os.path.exists(LOCK_FILE_PATH):
        QMessageBox.critical(None, "중복 실행 차단", "❌ 사줄게요가 이미 실행 중입니다.")
        sys.exit(0)
    with open(LOCK_FILE_PATH, "w", encoding="utf-8") as f:
        f.write("🔐 락파일 생성됨")
    atexit.register(lambda: os.remove(LOCK_FILE_PATH) if os.path.exists(LOCK_FILE_PATH) else None)
    window = SajulgeyoApp()
    window.show()
    sys.exit(app.exec_())